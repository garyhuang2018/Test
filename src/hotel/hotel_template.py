# encoding= utf-8
# __author__= gary

import sys
import pandas as pd
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QLabel, QCheckBox, QComboBox
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class TextToExcelApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.load_product_models()

    def initUI(self):
        self.setWindowTitle('一页纸调试模板')

        # Layout
        layout = QVBoxLayout()

        # Label
        self.label = QLabel('输入酒店名称:')
        layout.addWidget(self.label)

        # Text input and checkboxes layout
        input_layout = QHBoxLayout()

        # Text input
        self.text_input = QLineEdit(self)
        input_layout.addWidget(self.text_input)

        # Checkboxes
        self.checkbox1 = QCheckBox('展箱', self)
        self.checkbox2 = QCheckBox('样板间', self)
        self.checkbox3 = QCheckBox('批量', self)
        input_layout.addWidget(self.checkbox1)
        input_layout.addWidget(self.checkbox2)
        input_layout.addWidget(self.checkbox3)

        layout.addLayout(input_layout)

        # Product model dropdown
        self.product_model_dropdown = QComboBox(self)
        layout.addWidget(self.product_model_dropdown)

        # Button
        self.button = QPushButton('保存到Excel', self)
        self.button.clicked.connect(self.save_to_excel)
        layout.addWidget(self.button)

        # Set layout
        self.setLayout(layout)
        self.show()

    def load_product_models(self):
        # Read the Excel file without skipping rows
        df = pd.read_excel('data/酒店产品物料模型对照表.xls')

        # Detect the row containing the 'productModel' keyword
        product_model_row = df.apply(lambda row: row.astype(str).str.contains('productModel').any(), axis=1).idxmax()

        # Read the Excel file again, skipping the rows before the detected row
        df = pd.read_excel('data/酒店产品物料模型对照表.xls', skiprows=product_model_row+1)

        # Extract the productModel column
        product_models = df['productModel'].dropna().unique()

        # Populate the dropdown
        self.product_model_dropdown.addItems(product_models)

    def save_to_excel(self):
        user_input = self.text_input.text()
        checked_options = []

        if self.checkbox1.isChecked():
            checked_options.append(self.checkbox1.text())
        if self.checkbox2.isChecked():
            checked_options.append(self.checkbox2.text())
        if self.checkbox3.isChecked():
            checked_options.append(self.checkbox3.text())

        df = pd.DataFrame([user_input], columns=["UserInput"])
        df['CheckedOptions'] = ', '.join(checked_options)
        df.to_excel("output.xlsx", index=False)

        # Load the workbook and select the active worksheet
        wb = load_workbook("output.xlsx")
        ws = wb.active

        # Merge cells A1 to D1
        ws.merge_cells('A1:D1')

        # Set the value and alignment for the merged cell
        ws['A1'] = user_input
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Save the workbook
        wb.save("output.xlsx")

        self.label.setText('Text saved to output.xlsx')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = TextToExcelApp()
    sys.exit(app.exec_())