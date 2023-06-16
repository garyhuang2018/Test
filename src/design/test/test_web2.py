from selenium import webdriver
import time
import openpyxl
import pytest
from selenium.webdriver.common.by import By


class Test_open():

    def setup_class(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')

        self.driver = webdriver.Chrome(options=options)
        # self.driver = webdriver.Chrome()
        path = 'ip.xlsx'
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        web_address = ws.cell(row=1, column=2).value
        username = ws.cell(row=2, column=2).value
        password = ws.cell(row=3, column=2).value
        wb.close()
        self.driver.get("http://" + web_address + "/#/login")
        self.driver.maximize_window()
        time.sleep(3)
        #输入账号
        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[1]/div/div/div/form/div[1]/div/div[1]/input').send_keys(username)
        time.sleep(2)
        #输入密码
        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[1]/div/div/div/form/div[2]/div/div/input').send_keys(password)
        time.sleep(2)
        #点击登录
        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[1]/div/div/div/form/div[3]/div/button').click()
        time.sleep(3)
        #点击智能物联
        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[1]/div[1]/div[2]/ul/li[2]').click()
        time.sleep(4)

    def test_case1(self):
        #循环重启
        #读取ip表格
        path = 'ip.xlsx'
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        reboot_num = ws.cell(row=4, column=2).value
        wb.close()
        for i in range(0, reboot_num):
            for row in ws.iter_rows(values_only=True):
                ip_address = row[2]
                if ip_address is not None:

                    self.driver.refresh()
                    time.sleep(2)
                    try:
                        # 点击倒三角
                        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[2]/main/div[2]/section/div[1]/div/form/div/div[1]/div/div/div/div').click()
                        time.sleep(4)
                        # self.driver.find_element_by_xpath('/html/body/div[3]/div[1]/div[1]/ul')
                        # self.driver.find_elements(By.XPATH, '/html/body/div[4]/div[1]/div[1]/ul').click()
                        # 点击设备IP
                        self.driver.find_element(By.XPATH,'/html/body/div[2]/div[1]/div[1]/ul/li[3]').click()
                        # self.driver.find_element_by_xpath('/html/body/div[3]/div[1]/div[1]/ul/li[3]').click()
                        time.sleep(4)
                        #输入设备IP
                        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[2]/main/div[2]/section/div[1]/div/form/div/div[1]/div/input').send_keys(ip_address)
                        time.sleep(3)
                        #点击查询
                        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[2]/main/div[2]/section/div[1]/div/div[1]/button[3]').click()
                        time.sleep(3)
                        #点击详情
                        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[2]/main/div[2]/section/div[1]/div/div[2]/div[4]/div[2]/table/tbody/tr/td[13]/div/button[2]').click()
                        time.sleep(3)
                        #点击设备重启
                        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[2]/main/div[2]/section/div[1]/div[2]/form/div[2]/div/button[2]').click()
                        time.sleep(3)
                        # 点击确定重启
                        self.driver.find_element(By.XPATH, '/html/body/div[2]/div/div[3]/button[2]').click()
                        time.sleep(5)
                        #返回上一页
                        self.driver.find_element(By.XPATH,'//*[@id="app"]/div/div[2]/main/div[2]/section/div[1]/div[1]/div/i').click()
                        time.sleep(5)
                    except BaseException as e:
                        self.driver.refresh()
                        print('元素未定位到,页面刷新')
            time.sleep(20)
            print(i, "次循环执行成功")

    def teardown_class(self):
        self.driver.quit()
        print('执行成功')


if __name__ == '__main__':
    pytest.main(['-s','test_web2.py'])







