# encoding= utf-8
# __author__= gary
import openpyxl
import xlwt


def save_to_excel(dic):
    workbook = openpyxl.Workbook()
    sheet = workbook.active          # 获得一个的工作表
    sheet.title = dic.get('title')
    i = 1
    for key in dic.keys():
        print(key)
        sheet.cell(1, i, key)
        sheet.cell(2, i, dic[key])
        i = i + 1
        # worksheet.write( 1,1,dic[key])  # start from line 2 to write data to excel

    workbook.save("test-record.xlsx")
    workbook.close()


def open_excel():
    wb = openpyxl.load_workbook('test-record.xlsx')
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.coordinate, cell.value)


if __name__ == '__main__':
    dic = dict(title="KEN12", build_no="20230407", id="02")
    save_to_excel(dic)
    # open_excel()
